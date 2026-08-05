"""
market_brand_performance.py
===========================
Cross-Market Brand Performance & Competitor Intelligence Pipeline.

Objective:
    Extracts and normalizes brand metrics across BeautyByKat and all competitor 
    stores (Glowin, Sokostore, XBeauty, etc.) to generate a multi-tab Excel report:
    - Tab 1: Overall Market Performance (Consolidated cross-store brand summary).
    - Tab 2+: Individual store-level brand performance sheets (BeautyByKat + Competitors).

Databases:
    - Internal DB: beautybykat_inventory.db
    - Competitor DB: master_competitor.db

Output:
    Excel Report: Pipeline/Reports/output/Market_Brand_Performance_Report_YYYY-MM-DD.xlsx
"""

import sqlite3
import logging
import sys
import re
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# 1. LOGGING & DYNAMIC PATH RESOLUTION
# =====================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger("MarketBrandPerformance")

SCRIPT_DIR = Path(__file__).resolve().parent

def find_pipeline_root(start_dir: Path) -> Path:
    """Walks up parent directories to dynamically locate the pipeline root."""
    for parent in [start_dir] + list(start_dir.parents):
        if (parent / "orchestrate.py").exists() or (parent / "Data" / "databases").exists():
            return parent
    return start_dir.parents[1] if len(start_dir.parents) >= 2 else start_dir

PIPELINE_ROOT = find_pipeline_root(SCRIPT_DIR)

# Core Directories & Databases
DB_DIR = PIPELINE_ROOT / "Data" / "databases"
BBK_DB = DB_DIR / "beautybykat_inventory.db"
COMP_DB = DB_DIR / "master_competitor.db"

OUTPUT_DIR = PIPELINE_ROOT / "Reports" / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TODAY = datetime.now()
OUTPUT_FILE = OUTPUT_DIR / f"Market_Brand_Performance_Report_{TODAY.strftime('%Y-%m-%d')}.xlsx"


# =====================================================================
# 2. BRAND CLEANING & NORMALIZATION HELPERS
# =====================================================================
def clean_brand_name(name: str) -> str:
    """Normalizes brand titles across internal and competitor datasets."""
    if not name or pd.isna(name):
        return "UNKNOWN BRAND"
    
    clean = str(name).strip().upper()
    aliases = {
        "DR ALTHEA": "DR.ALTHEA",
        "PURITO SEOUL": "PURITO",
        "IUNIK": "IUNIK",
        "DR JART+": "DR.JART+",
        "DR CEURACLE": "DR.CEURACLE"
    }
    return aliases.get(clean, clean)


# =====================================================================
# 3. BEAUTYBYKAT DATA EXTRACTION MODULE
# =====================================================================
def get_bbk_brand_data(conn: sqlite3.Connection) -> pd.DataFrame:
    """Extracts store-level brand performance from BeautyByKat database."""
    query = """
        WITH matched_line_items AS (
            SELECT 
                oli.line_item_id,
                COALESCE(pv.upk_id, p_fallback.upk_id) AS upk_id,
                oli.quantity,
                (oli.quantity * oli.price) AS line_revenue
            FROM order_line_items oli
            LEFT JOIN product_variants pv ON oli.variant_id = pv.variant_id
            LEFT JOIN products p_fallback ON (oli.variant_id IS NULL OR oli.variant_id = '')
                  AND oli.raw_lineitem_name LIKE (p_fallback.product_title || '%')
        )
        SELECT 
            'BeautyByKat' AS Company,
            b.clean_name AS Raw_Brand,
            COUNT(DISTINCT p.upk_id) AS "Products Listed",
            COALESCE(SUM(mli.quantity), 0) AS "Units Sold",
            COALESCE(ROUND(SUM(mli.line_revenue), 3), 0.0) AS "Earnings (BHD)"
        FROM brands b
        LEFT JOIN products p ON b.brand_id = p.brand_id
        LEFT JOIN matched_line_items mli ON p.upk_id = mli.upk_id
        GROUP BY b.brand_id, b.clean_name;
    """
    df = pd.read_sql_query(query, conn)
    df["Brand Name"] = df["Raw_Brand"].apply(clean_brand_name)
    df.drop(columns=["Raw_Brand"], inplace=True)
    return df[["Company", "Brand Name", "Products Listed", "Units Sold", "Earnings (BHD)"]]


# =====================================================================
# 4. COMPETITOR DATA EXTRACTION MODULE
# =====================================================================
def get_competitor_brand_data(conn: sqlite3.Connection) -> pd.DataFrame:
    """
    Queries competitor stock snapshots to estimate units sold and revenue 
    via daily stock depletion analysis per competitor company.
    """
    # 1. Catalog Products per Company
    query_catalog = """
        SELECT 
            v.origin_company AS Company,
            b.canonical_name AS Raw_Brand,
            COUNT(DISTINCT v.upk_id) AS "Products Listed"
        FROM fact_store_variants v
        JOIN dim_unified_products p ON v.upk_id = p.upk_id
        JOIN dim_brands b ON p.brand_id = b.brand_id
        GROUP BY v.origin_company, b.canonical_name;
    """
    df_catalog = pd.read_sql_query(query_catalog, conn)

    # 2. Daily Stock Depletion Query
    query_snapshots = """
        SELECT 
            s.link_id,
            s.origin_company AS Company,
            b.canonical_name AS Raw_Brand,
            s.stock_quantity,
            COALESCE(NULLIF(s.price_bhd, 0), NULLIF(v.price_bhd, 0), 0.0) AS price_bhd,
            s.snapshot_date
        FROM fact_daily_stock_snapshots s
        JOIN fact_store_variants v ON s.link_id = v.link_id
        JOIN dim_unified_products p ON s.upk_id = p.upk_id
        JOIN dim_brands b ON p.brand_id = b.brand_id
        ORDER BY s.link_id, s.snapshot_date ASC;
    """
    df_snaps = pd.read_sql_query(query_snapshots, conn)

    if df_snaps.empty:
        df_catalog["Brand Name"] = df_catalog["Raw_Brand"].apply(clean_brand_name)
        df_catalog["Units Sold"] = 0
        df_catalog["Earnings (BHD)"] = 0.0
        return df_catalog[["Company", "Brand Name", "Products Listed", "Units Sold", "Earnings (BHD)"]]

    # Process stock depletions
    df_snaps["stock_quantity"] = pd.to_numeric(df_snaps["stock_quantity"], errors="coerce").fillna(0)
    df_snaps["price_bhd"] = pd.to_numeric(df_snaps["price_bhd"], errors="coerce").fillna(0.0)
    df_snaps["prev_stock"] = df_snaps.groupby("link_id")["stock_quantity"].shift(1)
    
    # Positive delta = items sold
    df_snaps["units_sold"] = (df_snaps["prev_stock"] - df_snaps["stock_quantity"]).apply(
        lambda x: int(x) if pd.notnull(x) and x > 0 else 0
    )
    df_snaps["revenue_bhd"] = df_snaps["units_sold"] * df_snaps["price_bhd"]

    # Aggregate by Company & Brand
    df_sales = df_snaps.groupby(["Company", "Raw_Brand"]).agg(
        **{"Units Sold": ("units_sold", "sum"), "Earnings (BHD)": ("revenue_bhd", "sum")}
    ).reset_index()

    # Merge Catalog Counts with Sales
    df_merged = pd.merge(df_catalog, df_sales, on=["Company", "Raw_Brand"], how="left")
    df_merged["Units Sold"] = df_merged["Units Sold"].fillna(0).astype(int)
    df_merged["Earnings (BHD)"] = df_merged["Earnings (BHD)"].fillna(0.0).round(3)
    df_merged["Brand Name"] = df_merged["Raw_Brand"].apply(clean_brand_name)

    # Normalize Company Name capitalization
    df_merged["Company"] = df_merged["Company"].str.strip().str.title()

    return df_merged[["Company", "Brand Name", "Products Listed", "Units Sold", "Earnings (BHD)"]]


# =====================================================================
# 5. MARKET CONSOLIDATION & OVERALL SUMMARY MODULE
# =====================================================================
def build_overall_market_summary(df_all: pd.DataFrame) -> pd.DataFrame:
    """Aggregates metrics across all companies into a single market view."""
    # Market Totals
    overall = df_all.groupby("Brand Name").agg(
        **{
            "Total Stores Selling": ("Company", "nunique"),
            "Market Products Listed": ("Products Listed", "sum"),
            "Total Market Units Sold": ("Units Sold", "sum"),
            "Total Market Earnings (BHD)": ("Earnings (BHD)", "sum")
        }
    ).reset_index()

    # Calculate Top Store per Brand
    top_stores = df_all.sort_values(
        by=["Brand Name", "Earnings (BHD)", "Units Sold"], 
        ascending=[True, False, False]
    ).groupby("Brand Name").first().reset_index()
    
    top_stores = top_stores[["Brand Name", "Company"]].rename(columns={"Company": "Top Store for Brand"})

    # Extract BeautyByKat Earnings for Market Share calculation
    bbk_data = df_all[df_all["Company"] == "BeautyByKat"][["Brand Name", "Earnings (BHD)"]].rename(
        columns={"Earnings (BHD)": "BBK Earnings (BHD)"}
    )

    # Merge Market Summaries
    summary = pd.merge(overall, top_stores, on="Brand Name", how="left")
    summary = pd.merge(summary, bbk_data, on="Brand Name", how="left")
    summary["BBK Earnings (BHD)"] = summary["BBK Earnings (BHD)"].fillna(0.0)

    # Compute BBK Market Share %
    summary["BBK Market Share (%)"] = (
        (summary["BBK Earnings (BHD)"] / summary["Total Market Earnings (BHD)"]) * 100
    ).fillna(0.0).round(2)

    # Final Column Formatting
    cols_order = [
        "Brand Name", 
        "Total Stores Selling", 
        "Market Products Listed", 
        "Total Market Units Sold", 
        "Total Market Earnings (BHD)", 
        "BBK Earnings (BHD)", 
        "BBK Market Share (%)", 
        "Top Store for Brand"
    ]
    
    return summary[cols_order].sort_values(by="Total Market Earnings (BHD)", ascending=False)


# =====================================================================
# 6. EXCEL STYLING MODULE
# =====================================================================
def style_excel_report(file_path: Path):
    """Applies professional formatting across all sheets in the output workbook."""
    wb = openpyxl.load_workbook(file_path)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(name="Segoe UI", size=11, bold=True, color="000000")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
    )
    total_border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        # Header Formatting
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data Row Formatting
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                col_name = str(ws.cell(row=1, column=col).value)

                if "Earnings" in col_name or "(BHD)" in col_name:
                    cell.number_format = '#,##0.000 "BHD"'
                    cell.alignment = Alignment(horizontal="right")
                elif "%" in col_name or "Share" in col_name:
                    cell.number_format = '0.00"%"'
                    cell.alignment = Alignment(horizontal="right")
                elif "Units" in col_name or "Products" in col_name or "Stores" in col_name:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # Summary Totals Row
        tot_row = max_row + 1
        ws.cell(row=tot_row, column=1, value="TOTAL")
        
        for col in range(1, max_col + 1):
            cell = ws.cell(row=tot_row, column=col)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = total_border
            col_name = str(ws.cell(row=1, column=col).value)
            col_letter = get_column_letter(col)

            if col == 1:
                cell.alignment = Alignment(horizontal="left")
            elif "Earnings" in col_name or "(BHD)" in col_name:
                cell.value = f"=SUM({col_letter}2:{col_letter}{max_row})"
                cell.number_format = '#,##0.000 "BHD"'
                cell.alignment = Alignment(horizontal="right")
            elif "Units" in col_name or "Products" in col_name:
                cell.value = f"=SUM({col_letter}2:{col_letter}{max_row})"
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    wb.save(file_path)


# =====================================================================
# 7. MAIN ORCHESTRATOR
# =====================================================================
def main():
    logger.info("Initializing Cross-Market Brand Performance Pipeline...")

    df_list = []

    # 1. Process BeautyByKat
    if BBK_DB.exists():
        logger.info(f"Connecting to Internal DB: {BBK_DB.name}")
        conn_bbk = sqlite3.connect(BBK_DB)
        try:
            df_bbk = get_bbk_brand_data(conn_bbk)
            df_list.append(df_bbk)
            logger.info(f"Retrieved {len(df_bbk)} brands for BeautyByKat.")
        finally:
            conn_bbk.close()
    else:
        logger.warning(f"Internal DB not found at: {BBK_DB}")

    # 2. Process Competitors
    if COMP_DB.exists():
        logger.info(f"Connecting to Competitor DB: {COMP_DB.name}")
        conn_comp = sqlite3.connect(COMP_DB)
        try:
            df_comp = get_competitor_brand_data(conn_comp)
            if not df_comp.empty:
                df_list.append(df_comp)
                logger.info(f"Retrieved competitor metrics across {df_comp['Company'].nunique()} companies.")
        finally:
            conn_comp.close()
    else:
        logger.warning(f"Competitor DB not found at: {COMP_DB}")

    if not df_list:
        logger.error("No data sources available. Pipeline aborted.")
        return

    # Combine all store datasets
    df_all_stores = pd.concat(df_list, ignore_index=True)

    # Build Market Summary
    df_market_summary = build_overall_market_summary(df_all_stores)

    # 3. Export to Excel Multi-Tab Report
    logger.info(f"Generating multi-tab Excel report at: {OUTPUT_FILE}")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        # Sheet 1: Market Overall Performance Summary
        df_market_summary.to_excel(writer, index=False, sheet_name="Overall Market Performance")

        # Sheet 2+: Company Breakdown Sheets
        companies = df_all_stores["Company"].unique()
        for company in sorted(companies):
            df_company = df_all_stores[df_all_stores["Company"] == company].drop(columns=["Company"])
            df_company = df_company.sort_values(by="Earnings (BHD)", ascending=False)
            
            # Sanitize sheet name for Excel standards (max 31 chars)
            safe_sheet = re.sub(r'[\\/*?:\[\]]', '', str(company))[:31]
            df_company.to_excel(writer, index=False, sheet_name=safe_sheet)

    # 4. Apply Visual Formatting
    style_excel_report(OUTPUT_FILE)
    logger.info(f"[SUCCESS] Multi-Tab Brand Report generated successfully at: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()