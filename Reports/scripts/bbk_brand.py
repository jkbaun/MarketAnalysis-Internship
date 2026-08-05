"""
bbk_brand_performance.py
========================
ETL Pipeline & BI Reporting Script for BeautyByKat.

Objective:
    Generates a brand performance report listing all brands BeautyByKat sells, 
    the total count of products for each brand, total units sold, and total 
    earnings (revenue) generated per brand.

Database:
    Internal DB: beautybykat_inventory.db

Outputs:
    Excel Report: Pipeline/Reports/output/BBK_Brand_Performance_Report_YYYY-MM-DD.xlsx
"""

import sqlite3
import logging
import sys
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# 1. LOGGING & PATH RESOLUTION
# =====================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger("BBKBrandReport")

SCRIPT_DIR = Path(__file__).resolve().parent

def find_pipeline_root(start_dir: Path) -> Path:
    """Walks up parent directories to dynamically locate the pipeline root."""
    for parent in [start_dir] + list(start_dir.parents):
        if (parent / "orchestrate.py").exists() or (parent / "Data" / "databases").exists():
            return parent
    return start_dir.parents[1] if len(start_dir.parents) >= 2 else start_dir

PIPELINE_ROOT = find_pipeline_root(SCRIPT_DIR)

# Core Directories & Files
DB_DIR = PIPELINE_ROOT / "Data" / "databases"
BBK_DB = DB_DIR / "beautybykat_inventory.db"

OUTPUT_DIR = PIPELINE_ROOT / "Reports" / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TODAY = datetime.now()
OUTPUT_FILE = OUTPUT_DIR / f"BBK_Brand_Performance_Report_{TODAY.strftime('%Y-%m-%d')}.xlsx"


# =====================================================================
# 2. DATA EXTRACTION MODULE
# =====================================================================
def get_brand_performance(conn: sqlite3.Connection) -> pd.DataFrame:
    """
    Queries the BeautyByKat database to summarize brand metrics.
    
    SQL Architecture Logic:
    - Uses a Common Table Expression (CTE 'matched_line_items') to attribute 
      each order line item to a single UPK product ID (via variant_id or title match).
    - Left joins the 'brands' catalog table to ensure ALL brands are included in 
      the final report, even if they have 0 products or 0 sales.
    - Aggregates distinct product counts, total units sold, and total revenue.
    """
    query = """
        WITH matched_line_items AS (
            SELECT 
                oli.line_item_id,
                COALESCE(pv.upk_id, p_fallback.upk_id) AS upk_id,
                oli.quantity,
                (oli.quantity * oli.price) AS line_revenue
            FROM order_line_items oli
            LEFT JOIN product_variants pv 
                   ON oli.variant_id = pv.variant_id
            LEFT JOIN products p_fallback 
                   ON (oli.variant_id IS NULL OR oli.variant_id = '')
                  AND oli.raw_lineitem_name LIKE (p_fallback.product_title || '%')
        )
        SELECT 
            b.clean_name AS "Brand Name",
            COUNT(DISTINCT p.upk_id) AS "Total Products Listed",
            COALESCE(SUM(mli.quantity), 0) AS "Total Units Sold",
            COALESCE(ROUND(SUM(mli.line_revenue), 3), 0.0) AS "Total Earnings (BHD)"
        FROM brands b
        LEFT JOIN products p 
               ON b.brand_id = p.brand_id
        LEFT JOIN matched_line_items mli 
               ON p.upk_id = mli.upk_id
        GROUP BY b.brand_id, b.clean_name
        ORDER BY "Total Earnings (BHD)" DESC, "Total Products Listed" DESC;
    """
    
    df = pd.read_sql_query(query, conn)
    
    # Type safety conversions
    df["Total Products Listed"] = pd.to_numeric(df["Total Products Listed"], errors="coerce").fillna(0).astype(int)
    df["Total Units Sold"] = pd.to_numeric(df["Total Units Sold"], errors="coerce").fillna(0).astype(int)
    df["Total Earnings (BHD)"] = pd.to_numeric(df["Total Earnings (BHD)"], errors="coerce").fillna(0.0)
    
    return df


# =====================================================================
# 3. EXCEL STYLING & REPORT FORMATTING MODULE
# =====================================================================
def style_excel_report(file_path: Path):
    """
    Applies professional visual styling to the output Excel workbook:
    - Navy Blue header (#1F4E78) with bold white text.
    - BHD Currency formatting (#,##0.000 "BHD") for earnings.
    - Integer formatting (#,##0) for products and units sold.
    - Executive Summary Total row at the bottom.
    - Auto-adjust column widths for readability.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.title = "Brand Performance"

    # Styling Elements
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(name="Segoe UI", size=11, bold=True, color="000000")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    top_thick_bottom_double = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    max_row = ws.max_row
    max_col = ws.max_column

    # 1. Format Headers
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 2. Format Data Rows
    for row in range(2, max_row + 1):
        # Brand Name (Col 1)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=1).border = thin_border
        
        # Total Products (Col 2)
        cell_prod = ws.cell(row=row, column=2)
        cell_prod.number_format = '#,##0'
        cell_prod.alignment = Alignment(horizontal="right", vertical="center")
        cell_prod.border = thin_border
        
        # Total Units Sold (Col 3)
        cell_units = ws.cell(row=row, column=3)
        cell_units.number_format = '#,##0'
        cell_units.alignment = Alignment(horizontal="right", vertical="center")
        cell_units.border = thin_border

        # Total Earnings (Col 4)
        cell_earn = ws.cell(row=row, column=4)
        cell_earn.number_format = '#,##0.000 "BHD"'
        cell_earn.alignment = Alignment(horizontal="right", vertical="center")
        cell_earn.border = thin_border

    # 3. Add Summary Totals Row
    summary_row = max_row + 1
    ws.cell(row=summary_row, column=1, value="TOTAL")
    ws.cell(row=summary_row, column=2, value=f"=SUM(B2:B{max_row})")
    ws.cell(row=summary_row, column=3, value=f"=SUM(C2:C{max_row})")
    ws.cell(row=summary_row, column=4, value=f"=SUM(D2:D{max_row})")

    # Style Summary Totals Row
    for col in range(1, max_col + 1):
        cell = ws.cell(row=summary_row, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = top_thick_bottom_double
        
        if col == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col in [2, 3]:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col == 4:
            cell.number_format = '#,##0.000 "BHD"'
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # 4. Auto-Fit Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

    wb.save(file_path)


# =====================================================================
# 4. MAIN ORCHESTRATOR
# =====================================================================
def main():
    logger.info("Starting BeautyByKat Brand Performance Pipeline...")

    if not BBK_DB.exists():
        logger.error(f"Internal Database missing at path: {BBK_DB}")
        return

    logger.info(f"Connecting to Internal Database: {BBK_DB.name}")
    conn = None
    try:
        conn = sqlite3.connect(BBK_DB)
        df_brand = get_brand_performance(conn)
        logger.info(f"Successfully processed metrics for {len(df_brand)} brands.")

        # Export raw DataFrame to Excel
        logger.info(f"Writing Excel report to: {OUTPUT_FILE}")
        with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
            df_brand.to_excel(writer, index=False, sheet_name="Brand Performance")

        # Apply OpenPyXL Visual Formatting
        style_excel_report(OUTPUT_FILE)
        logger.info(f"[SUCCESS] BeautyByKat Brand Performance Report generated successfully at: {OUTPUT_FILE}")

    except Exception as e:
        logger.error(f"Pipeline execution failed: {str(e)}", exc_info=True)
    finally:
        if conn:
            conn.close()
            logger.info("Database connection closed.")


if __name__ == "__main__":
    main()