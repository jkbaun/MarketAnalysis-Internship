import sqlite3
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill

# =====================================================================
# CONFIGURATION & PATHS
# =====================================================================
DB_DIR = Path("/home/boredom-speaking/Desktop/JulyInternship/Pipeline/Data/databases")
COMP_DB = DB_DIR / "master_competitor.db"

OUTPUT_DIR = Path("/home/boredom-speaking/Desktop/JulyInternship/Pipeline/Reports/output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
TODAY = datetime.now()
OUTPUT_FILE = OUTPUT_DIR / f"Market_Analytics_Report_{TODAY.strftime('%Y-%m-%d')}.xlsx"

ANALYSIS_DAYS = 7

def get_snapshot_data(conn):
    cutoff_date = (TODAY - timedelta(days=ANALYSIS_DAYS)).strftime('%Y-%m-%d')
    query = """
        SELECT s.link_id, s.upk_id, s.origin_company, s.snapshot_date, s.stock_quantity, s.price_bhd, p.consensus_title, p.extracted_spec, b.canonical_name AS brand_name
        FROM fact_daily_stock_snapshots s
        JOIN dim_unified_products p ON s.upk_id = p.upk_id
        JOIN dim_brands b ON p.brand_id = b.brand_id
        WHERE DATE(s.snapshot_date) >= ?
        ORDER BY s.link_id, s.snapshot_date ASC
    """
    df = pd.read_sql_query(query, conn, params=(cutoff_date,))
    if df.empty: return pd.DataFrame()
    
    df['snapshot_date'] = pd.to_datetime(df['snapshot_date'])
    df['prev_stock'] = df.groupby('link_id')['stock_quantity'].shift(1)
    df['stock_diff'] = df['prev_stock'] - df['stock_quantity']
    df['units_sold'] = df['stock_diff'].apply(lambda x: int(x) if (pd.notnull(x) and x > 0) else 0)
    df['revenue_bhd'] = df['units_sold'] * df['price_bhd']
    return df

def get_price_matrix(conn):
    query = """
        WITH LatestSnapshots AS (
            SELECT s.upk_id, s.origin_company, s.price_bhd, s.snapshot_date, p.consensus_title, p.extracted_spec, b.canonical_name AS brand_name,
            ROW_NUMBER() OVER(PARTITION BY s.upk_id, s.origin_company ORDER BY s.snapshot_date DESC) as rn
            FROM fact_daily_stock_snapshots s
            JOIN dim_unified_products p ON s.upk_id = p.upk_id
            JOIN dim_brands b ON p.brand_id = b.brand_id WHERE s.price_bhd > 0
        ), SharedProducts AS (
            SELECT upk_id FROM LatestSnapshots WHERE rn = 1 GROUP BY upk_id HAVING COUNT(DISTINCT origin_company) > 1
        )
        SELECT ls.upk_id, ls.brand_name, ls.consensus_title, ls.extracted_spec, ls.origin_company, ls.price_bhd
        FROM LatestSnapshots ls JOIN SharedProducts sp ON ls.upk_id = sp.upk_id WHERE ls.rn = 1
    """
    raw_df = pd.read_sql_query(query, conn)
    if raw_df.empty: return pd.DataFrame(), []
    
    pivot = raw_df.pivot_table(index=['upk_id', 'brand_name', 'consensus_title', 'extracted_spec'], columns='origin_company', values='price_bhd', aggfunc='first').reset_index()
    company_cols = [c for c in pivot.columns if c not in ['upk_id', 'brand_name', 'consensus_title', 'extracted_spec']]
    pivot['Min Price (BHD)'] = pivot[company_cols].min(axis=1)
    pivot['Max Price (BHD)'] = pivot[company_cols].max(axis=1)
    pivot['Price Difference (BHD)'] = pivot['Max Price (BHD)'] - pivot['Min Price (BHD)']
    pivot['Price Variance (%)'] = (pivot['Price Difference (BHD)'] / pivot['Min Price (BHD)']) * 100
    pivot.rename(columns={'upk_id': 'UPK ID', 'brand_name': 'Brand', 'consensus_title': 'Product Title', 'extracted_spec': 'Spec / Volume'}, inplace=True)
    return pivot.sort_values(by='Price Variance (%)', ascending=False), company_cols

def get_shared_unique(conn):
    query = """
        SELECT u.upk_id, b.canonical_name AS root_brand_name, u.consensus_title, u.extracted_spec, u.canonical_barcode, u.canonical_sku, v.origin_company, v.price_bhd
        FROM dim_unified_products u
        LEFT JOIN dim_brands b ON u.brand_id = b.brand_id
        JOIN fact_store_variants v ON u.upk_id = v.upk_id
    """
    df_raw = pd.read_sql_query(query, conn)
    if df_raw.empty: return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    company_summary = df_raw.groupby(['upk_id', 'origin_company']).agg({'price_bhd': 'min'}).reset_index()
    pivot_bhd = company_summary.pivot(index='upk_id', columns='origin_company', values='price_bhd')
    pivot_bhd.columns = [f"price_bhd_{col}" for col in pivot_bhd.columns]
    
    upk_metadata = df_raw.groupby('upk_id').agg({'root_brand_name': 'first', 'consensus_title': 'first', 'extracted_spec': 'first', 'canonical_barcode': 'first', 'canonical_sku': 'first'})
    merged = upk_metadata.join(pivot_bhd)
    
    bhd_cols = [c for c in merged.columns if c.startswith('price_bhd_')]
    merged['store_count'] = merged[bhd_cols].notna().sum(axis=1)
    merged['selling_stores'] = merged.apply(lambda row: ", ".join([col.replace('price_bhd_', '') for col in bhd_cols if pd.notna(row[col])]), axis=1)
    merged['product_type'] = merged['store_count'].apply(lambda x: 'Shared' if x > 1 else 'Unique')
    merged['min_price_bhd'] = merged[bhd_cols].min(axis=1)
    merged['max_price_bhd'] = merged[bhd_cols].max(axis=1)
    merged['price_spread_bhd'] = round(merged['max_price_bhd'] - merged['min_price_bhd'], 3)
    
    final_cols = ['root_brand_name', 'consensus_title', 'extracted_spec', 'canonical_barcode', 'canonical_sku', 'product_type', 'store_count', 'selling_stores', 'min_price_bhd', 'max_price_bhd', 'price_spread_bhd'] + sorted(bhd_cols)
    df_final = merged.reset_index()[['upk_id'] + final_cols]
    
    return df_final, df_final[df_final['product_type'] == 'Shared'].copy(), df_final[df_final['product_type'] == 'Unique'].copy()

def style_excel(file_path, comp_cols):
    wb = openpyxl.load_workbook(file_path)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            
        if sheet_name == "Price Matrix":
            headers = [str(ws.cell(row=1, column=col).value) for col in range(1, ws.max_column + 1)]
            comp_col_indices = [headers.index(c) + 1 for c in comp_cols if c in headers]
            green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            
            for row in range(2, ws.max_row + 1):
                store_prices = {idx: ws.cell(row=row, column=idx).value for idx in comp_col_indices if isinstance(ws.cell(row=row, column=idx).value, (int, float)) and ws.cell(row=row, column=idx).value > 0}
                min_p = min(store_prices.values()) if store_prices else None
                max_p = max(store_prices.values()) if store_prices else None
                
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if col in comp_col_indices and cell.value is not None:
                        cell.number_format = '#,##0.000 "BHD"'
                        if min_p and max_p and min_p != max_p:
                            if cell.value == min_p: cell.fill = green_fill
                            elif cell.value == max_p: cell.fill = red_fill
    wb.save(file_path)

def main():
    if not COMP_DB.exists():
        print("[-] Competitor DB not found!")
        return
        
    conn = sqlite3.connect(COMP_DB)
    df_snapshots = get_snapshot_data(conn)
    df_price_matrix, comp_cols = get_price_matrix(conn)
    df_all, df_shared, df_unique = get_shared_unique(conn)
    conn.close()

    company_perf = pd.DataFrame()
    if not df_snapshots.empty:
        company_perf = df_snapshots.groupby('origin_company').agg(total_units_sold=('units_sold', 'sum'), total_revenue_bhd=('revenue_bhd', 'sum')).reset_index().sort_values(by='total_revenue_bhd', ascending=False)
    
    print(f"[+] Writing Market Analytics to {OUTPUT_FILE}...")
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df_price_matrix.to_excel(writer, sheet_name="Price Matrix", index=False)
        df_all.to_excel(writer, sheet_name="All Competitor Products", index=False)
        df_shared.to_excel(writer, sheet_name="Shared Products", index=False)
        df_unique.to_excel(writer, sheet_name="Unique Products", index=False)
        if not company_perf.empty:
            company_perf.to_excel(writer, sheet_name="Market Share (Est)", index=False)
        
    style_excel(OUTPUT_FILE, comp_cols)
    print(f"[SUCCESS] Market Analytics Report generated at {OUTPUT_FILE}")

if __name__ == "__main__":
    main()